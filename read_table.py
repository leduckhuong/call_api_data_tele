from openpyxl import load_workbook
import csv
from datetime import datetime
from check_custom_mail import check_custom_mail

# CUSTOMER_NAME BIRTH_DATE PHONE EMAIL PASSWORD
def read_table_xlsx(path):
    workbook = load_workbook(filename=path, data_only=True)
    sheet = workbook.active  # Hoặc bạn có thể chỉ định sheet cụ thể

    columns = [cell.value for cell in sheet[1]]

    require_columns = ['CUSTOMER_NAME', 'BIRTH_DATE', 'PHONE', 'EMAIL', 'PASSWORD']
    valid_columns = [col for col in columns if col in require_columns]

    if valid_columns:
    
        for row in sheet.iter_rows(min_row=2, values_only=True): 
            row_data = {col: row[i] for i, col in enumerate(columns) if col in valid_columns}

            if row_data['EMAIL'] is not None:
                if check_custom_mail(row_data['EMAIL']):
                    if 'BIRTH_DATE' in row_data and isinstance(row_data['BIRTH_DATE'], datetime):
                        row_data['BIRTH_DATE'] = row_data['BIRTH_DATE'].strftime('%d/%m/%Y')
                    print(row_data)

            
            
    else:
        print('Không có cột nào trong danh sách yêu cầu tồn tại trong file.')

def read_table_csv(path):
    with open(path, mode='r') as file:
        reader = csv.reader(file)
        for row in reader:
            print(row)

def main():
    path = '/mnt/500GB/downloads/data/Vietnam.xlsx'
    read_table_xlsx(path)

if __name__ == '__main__':
    main()
